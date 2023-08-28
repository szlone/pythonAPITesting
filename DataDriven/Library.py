import jsonpath
import requests
import json
import openpyxl


class Common:

    def __init__(self, FileNamePath, SheetName):
        global wk
        global sh
        wk = openpyxl.load_workbook('E:/MyCourses/RestAPI/pythonProject/StudentMS/StudentData.xlsx')
        sh = wk['sheet1']

    def fetch_row_count(self):
        rows = sh.max_row
        return rows

    def fetch_column_count(self):
        columns = sh.max_column
        return columns

    def fetch_key_names(self):
        c = sh.max_column
        li = []
        for i in range(1, c + 1):  # note first row is the header plus have to add 1 to include the last row
            cell = sh.cell(row=1, column=1)
            li.insert(i-1, cell.value)
        return li

    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sh.max_column
        for i in range(1, c+1):
            cell = sh.cell(row=rowNumber, column=1)
            jsonRequest[keyList[i-1]] = cell.value
            return jsonRequest





