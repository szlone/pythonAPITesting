import jsonpath
import requests
import json
import openpyxl

# adding all the tables at once....
def test_add_multiple_students():
    # API details
    API_URL = "https://thetestingworldapi.com/api/studentsDetails"
    f = open('E:/MyCourses/RestAPI/pythonProject/StudentMS/CreateNewStudent.json', 'r')
    json_request = json.loads(f.read())

    # Excel Code
    wk = openpyxl.load_workbook('E:/MyCourses/RestAPI/pythonProject/StudentMS/StudentData.xlsx')
    sh = wk['sheet1']
    rows = sh.max_row
    for i in range(2, rows+1):  # note first row is the header plus have to add 1 to include the last row
        cell_first_name = sh.cell(row=1, column=1)
        cell_middle_name = sh.cell(row=1, column=2)
        cell_last_name = sh.cell(row=1, column=3)
        cell_date_of_birth = sh.cell(row=1, column=4)
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(API_URL, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201

    # json_request = json.loads(f.read())
    # response = requests.post(API_URL, json_request)
    # id = jsonpath.jsonpath(response.json(), 'id')
    # print(response.text)
    # print(id[0])
    # print(response.status_code)
    # assert response.status_code == 201
